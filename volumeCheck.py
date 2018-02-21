"""
Client Volume Monitor

"""

import numpy as np
import pandas as pd
from pandas import TimeGrouper
import pyodbc
from pandas import ExcelWriter
from datetime import datetime as dt
import calendar
from matplotlib import pyplot as plt
from functools import lru_cache
import xlsxwriter
import openpyxl
from openpyxl import workbook, worksheet, load_workbook
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.compat import range
from openpyxl.utils import get_column_interval

# Declarations
y = dt.now().year
m = dt.now().month
d = dt.now().day
currMonthID = y * 100 + m
currMonthDays = calendar.monthrange(y, m)
mtd = (currMonthDays[1] / d)

# SQL
conn = pyodbc.connect(r'DRIVER={ODBC Driver 13 for SQL Server};'
                      r'SERVER=businteldw.stratose.com,1565;'
                      r'DATABASE=CAIDataWarehouse;'
                      r'Trusted_Connection=yes')

sql = '''
    select dc.ClientParentNameShort
        , Grouped = 
        case dpr.Product
            when 'Dental' then 'Dental'
            when 'Workers Comp' then 'WC'
            else 'Medical'
        end
        --, dpr.Product
        , dd.DateMonthID 
        , sum(fc.ClaimCount) Claims
        , sum(fc.CMAllowed) Charges
            
    from FactClaim fc
        join DimDate dd on fc.dimdatereceivedkey = dd.dimdatekey
        join dimclient dc on fc.dimclientkey = dc.dimclientkey
        join dimclaimeligible dce on fc.dimclaimeligiblekey = dce.dimclaimeligiblekey
        join dimdiscountmethod ddm on fc.dimdiscountmethodkey = ddm.dimdiscountmethodkey
        join dimprovider dp on fc.dimproviderkey = dp.dimproviderkey
        join dimservicetypecategory dstc on fc.dimservicetypecategorykey = dstc.dimservicetypecategorykey
        join dimnetwork dn on fc.dimnetworkkey = dn.dimnetworkkey
        join dimproduct dpr on fc.dimproductkey = dpr.dimproductkey
        join dimclaimtype dct on fc.dimclaimtypekey = dct.dimclaimtypekey
        join DimClaimStatus dcs on fc.DimClaimStatusKey = dcs.DimClaimStatusKey
    
    where 
        dce.ClaimEligible = 'Eligible'
            and dd.DateDay between (convert(date, getdate() - 70)) and (convert(date, getdate()))
              
    group by dc.ClientParentNameShort
        , dpr.Product
        , dd.DateMonthID
    
    order by dc.ClientParentNameShort
        , dpr.Product
        , dd.DateMonthID

'''

df1 = pd.read_sql(sql, conn)
conn.close()

# Calcs
df = df1.set_index(['ClientParentNameShort', 'Grouped', 'DateMonthID'])
df['claimsMTD'] = df['Claims'] * mtd
df['chargesMTD'] = df['Charges'] * mtd
df['lagClaim'] = df['Claims'].shift(1)
df['lagCharge'] = df['Charges'].shift(1)
df['diffClaim'] = df['claimsMTD'] - df['lagClaim']
df['diffCharge'] = df['chargesMTD'] - df['lagCharge']
df.dropna(inplace=True)
df.reset_index(inplace=True)
df = df[df['DateMonthID'] == currMonthID]
df['claims%Lag'] = df['claimsMTD'] / df['lagClaim'] - 1
df['charges%Lag'] = df['chargesMTD'] / df['lagCharge'] - 1

dfFlag = df[(df['charges%Lag'] >= .25) | (df['charges%Lag'] <= -.25)]

dfMed = dfFlag[dfFlag.Grouped == "Medical"]
dfDent = dfFlag[dfFlag.Grouped == "Dental"]
dfWC = dfFlag[dfFlag.Grouped == "WC"]

# Write results to excel
writer = pd.ExcelWriter(r'C:\Users\pallen\Documents\VolumeCheck.xlsx', engine='xlsxwriter')

dfMed.to_excel(writer, 'Medical', startrow=1, startcol=1, index=False)
dfDent.to_excel(writer, 'Dental', startrow=1, startcol=1, index=False)
dfWC.to_excel(writer, 'WC', startrow=1, startcol=1, index=False)

writer.save()


def style():
    wb = load_workbook(r'C:\Users\pallen\Documents\VolumeCheck.xlsx')

    head = NamedStyle(name="head")
    thin = Side(border_style="thin", color="000000")
    head.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    head.fill = PatternFill(fill_type="solid", start_color="662766", end_color="662766")
    head.font = Font(color="FFFFFF")
    head.al = Alignment(horizontal="center", vertical="center")

    ws = wb.get_sheet_by_name('Medical')

    for column in range(2, 15):
        ws.cell(row=2, column=column).style = head
        ws.column_dimensions.height = 20

    ws = wb.get_sheet_by_name('Dental')

    for column in range(2, 15):
        ws.cell(row=2, column=column).style = head

    ws = wb.get_sheet_by_name('WC')

    for column in range(2, 15):
        ws.cell(row=2, column=column).style = head

    wb.save(r'C:\Users\pallen\Documents\VolumeCheck.xlsx')

    return 0



